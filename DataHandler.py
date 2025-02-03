import os
import re
import openpyxl
import pandas as pd

class DataHandler:
    def __init__(self):
        self.download_folder = '/Users/claudio_hurtado/Downloads'
        self.pattern_cuenta_corriente = re.compile(r"Movimientos Nacionales de Cuenta Corriente.*\.xlsx", re.IGNORECASE)
        self.pattern_tarjeta_credito = re.compile(r"Movimientos Nacionales de Tarjeta de Crédito.*\.xlsx", re.IGNORECASE)

    def get_matching_files(self, pattern):
        all_files = os.listdir(self.download_folder)
        return [file for file in all_files if pattern.match(file)]

    def find_cell_with_value(self, sheet, value):
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value == value:
                    return cell.coordinate
        return None

    def select_data(self, sheet, start_cell):
        data = []

        for row in sheet.iter_rows(min_row=start_cell.row, min_col=start_cell.column, values_only=True):
            data.append(row)

        df = pd.DataFrame(data[1:], columns=data[0])
        df = df.drop(df.index[-1])
        return df

    def process_save_cuenta_corriente(self, df):
        df.rename(columns={'Monto $': 'Monto'}, inplace=True)
        df['Monto'] = df['Monto'].fillna(0).map(int)
        df['Detalle'] = (df['Detalle'].str.replace('Cargo por ', '')
                                      .str.replace('Abono por ', '')
                                      .str.replace(",", ""))

        for index, row in df.iterrows():
            detalle_value = row['Detalle']
            match = re.search(r'\d{2}/\d{2}/\d{4}', detalle_value)
            if match:
                start_index = match.start()
                df.at[index, 'Fecha'] = detalle_value[start_index:start_index+10]
                df.at[index, 'Detalle'] = detalle_value[:start_index-4]

        df = df[['Fecha', 'Categoria', 'Detalle', 'Monto']]
        df.to_csv("results/results_cc.csv", index=False)

    def process_save_tarjeta_credito(self, df):
        df.rename(columns={'Monto $': 'Monto'}, inplace=True)
        df['Categoria'] = 'Cargo'
        df = df[['Fecha', 'Categoria', 'Detalle', 'Monto']]
        df = df.iloc[:-1, :]
        df.to_csv("results/results_tc.csv", index=False)


    def process_latest_file(self, pattern):
        matching_files = self.get_matching_files(pattern)
        if matching_files:
            newest_file = max(matching_files, key=lambda file: os.path.getctime(os.path.join(self.download_folder, file)))
            newest_file_path = os.path.join(self.download_folder, newest_file)
            print("Newest file:", newest_file_path)
            workbook = openpyxl.load_workbook(newest_file_path)
            sheet = workbook.worksheets[0]

            cell_location = self.find_cell_with_value(sheet, "Fecha")
            if cell_location:
                print(f'Cell with value "Fecha" found at: {cell_location}')
                start_cell = sheet[cell_location]
                df_raw = self.select_data(sheet, start_cell)

                if 'corriente' in pattern.pattern.lower():
                    self.process_save_cuenta_corriente(df_raw)
                elif 'crédito' in pattern.pattern.lower():
                    self.process_save_tarjeta_credito(df_raw)
                else:
                    raise ValueError("Invalid pattern.")

            else:
                print('Cell with value "Fecha" not found.')
        else:
            print("No matching files found.")

    def delete_old_files(self, pattern):
        matching_files = self.get_matching_files(pattern)
        if matching_files:
            newest_file = max(matching_files, key=lambda file: os.path.getctime(os.path.join(self.download_folder, file)))
            for file in matching_files:
                if file != newest_file:
                    os.remove(os.path.join(self.download_folder, file))
                    print(f"Deleted file: {file}")

    
