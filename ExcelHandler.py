import os
import re
import openpyxl
import pandas as pd

class ExcelHandler:
    def __init__(self, download_folder):
        self.download_folder = download_folder
        self.pattern = re.compile(r"Movimientos Nacionales de Cuenta Corriente.*\.xlsx", re.IGNORECASE)

    def get_matching_files(self):
        all_files = os.listdir(self.download_folder)
        return [file for file in all_files if self.pattern.match(file)]

    def find_cell_with_value(self, sheet, value):
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value == value:
                    return cell.coordinate
        return None

    def select_and_save_data(self, file_path, start_cell):
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        data = []

        for row in sheet.iter_rows(min_row=start_cell.row, min_col=start_cell.column, values_only=True):
            data.append(row)

        df = pd.DataFrame(data[1:], columns=data[0])
        df = df.drop(df.index[-1])
        df.rename(columns={'Monto $': 'Monto'}, inplace=True)
        df['Monto'] = df['Monto'].fillna(0).map(int)
        df['Detalle'] = df['Detalle'].str.replace('Cargo por ', '').replace('Abono por ', '').replace(',', '')

        for index, row in df.iterrows():
            detalle_value = row['Detalle']
            match = re.search(r'\d{2}/\d{2}/\d{4}', detalle_value)
            if match:
                start_index = match.start()
                df.at[index, 'Fecha'] = detalle_value[start_index:start_index+10]
                df.at[index, 'Detalle'] = detalle_value[:start_index-4]

        df = df[['Fecha', 'Categoria', 'Detalle', 'Monto']]
        df.to_csv("results.csv", index=False)

    def process_latest_file(self):
        matching_files = self.get_matching_files()
        if matching_files:
            newest_file = max(matching_files, key=lambda file: os.path.getctime(os.path.join(self.download_folder, file)))
            newest_file_path = os.path.join(self.download_folder, newest_file)
            print("Newest file:", newest_file_path)
            workbook = openpyxl.load_workbook(newest_file_path)
            sheet = workbook.worksheets[0]

            cell_location = self.find_cell_with_value(sheet, "Fecha")
            if cell_location:
                print(f'Cell with value "Fecha" found at: {cell_location}')
                start_cell = sheet[cell_location]
                self.select_and_save_data(newest_file_path, start_cell)
            else:
                print('Cell with value "Fecha" not found.')
        else:
            print("No matching files found.")

# Example usage
if __name__ == "__main__":
    download_folder = "/Users/claudio_hurtado/Downloads"  # Update this path to your download folder
    handler = ExcelHandler(download_folder)
    handler.process_latest_file()
